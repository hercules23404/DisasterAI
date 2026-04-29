#!/usr/bin/env python3
"""
Create a formatted Excel file for Literature Survey Table
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Data for the literature survey table
data = {
    'No.': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    'Paper Title': [
        'An Introduction to Centralized Training for Decentralized Execution in Cooperative Multi-Agent Reinforcement Learning',
        'Multi-Agent Reinforcement Learning with Hierarchical Coordination for Emergency Responder Stationing',
        'Shortest Path Planning and Dynamic Rescue Forces Dispatching for Urban Flood Disasters',
        'Coupled GPU-Based Modeling of Dynamic-Wave Flow and Solute Transport in Floods with Cellular Automata Framework',
        'Online Algorithms for Ambulance Routing in Disaster Response with Time-Varying Victim Conditions',
        'Evaluating the Impact of Digital Elevation Models on Urban Flood Modeling',
        'Supply-Demand Mismatch Causes Substantial Deterioration in Prehospital Emergency Medical Service Under Disasters',
        'Fast Prediction of Urban Flooding Water Depth Based on CNN-LSTM',
        'Pre-Positioning Facility Location and Resource Allocation in Humanitarian Relief Operations Considering Deprivation Costs',
        'Cooperative Reward Shaping for Multi-Agent Pathfinding'
    ],
    'Reference': ['[5]', '[3]', '[14]', '[12]', '[16]', '[18]', '[29]', '[9]', '[20]', '[32]'],
    'Year': [2024, 2024, 2025, 2024, 2024, 2024, 2025, 2023, 2021, 2024],
    'Venue': [
        'arXiv preprint',
        'ICML 2024',
        'Scientific Reports (Nature)',
        'Natural Hazards and Earth System Sciences',
        'OR Spectrum (Springer)',
        'Water Resources Management (Springer)',
        'Communications Engineering (Nature)',
        'Water (MDPI)',
        'Sustainability (MDPI)',
        'arXiv preprint'
    ],
    'How We Used This Paper': [
        'Core Framework Foundation: Provides the theoretical foundation for our QMIX implementation. Validates our CTDE paradigm. Justifies why agents can train with global information but execute with only local observations. Cited in Section II-C and III-F.',
        'Highest Prestige Validation: ICML 2024 publication validates our problem domain at the highest academic level. Provides empirical evidence (5-second response time reduction). Cited in Section II-C and V-E to justify MARL approach.',
        'Direct Methodological Parallel: Most directly relevant paper to our work. Uses customized A* with Hungarian-style optimization for flood rescue. Validates our Hungarian algorithm choice. Cited extensively in Sections II-B, III-E, and V.',
        'Flood Simulation Alternative: Represents the CA approach we deliberately avoided. Validates that CA can provide real-time flood simulation with GPU. Cited in Sections II-A and VI to contrast with our D8 approach.',
        'Dynamic Dispatch Validation: Directly addresses time-varying victim conditions and stochastic travel times. Validates our dynamic recomputation approach. Cited in Sections I, II-B, III-E, V-B, and VII.',
        'DEM Resolution Justification: Comprehensive assessment of SRTM-30m and NASADEM for flood modeling. Validates our choice of SRTM 30m resolution DEM. Cited in Sections II-A, III-A, and VI.',
        'Problem Motivation & Risk Scoring: Demonstrates real-world consequences of poor resource allocation during disasters. Validates importance of supply-demand balance. Cited in Sections I, III-D, and VII.',
        'ML-Based Flood Prediction Alternative: Represents the deep learning approach (alternative to our D8 method). Achieves <10 second prediction time. Cited in Sections II-A and V-C to contrast ML vs. physics-based approaches.',
        'MCLP Pre-Positioning Foundation: Provides theoretical foundation for our MCLP-based unit pre-positioning. Validates importance of deprivation costs. Cited in Sections II-D, III-E, IV-A, and V-E.',
        'Reward Function Design: Provides theoretical foundation for our cooperative reward shaping approach. Validates that aligning rewards with global goals reduces agent selfishness. Cited in Sections II-E, III-F, and VII.'
    ],
    'Limitations Stated': [
        'Requires extensive training episodes (500 episodes). Assumes agents can access global state during training. Communication overhead during centralized training. May not scale to >50 agents.',
        'Uses hierarchical MARL (more complex than our Hungarian approach). Requires 3 orders of magnitude more computation. Tested only on ambulance repositioning, not flood scenarios. No predictive flood modeling.',
        'Does not incorporate N-step predictive flood simulation. Uses static flood maps rather than dynamic propagation. No comparison with MARL frameworks. Limited to small-scale scenarios.',
        'Requires GPU hardware for real-time performance. CA produces less topographically accurate flow than priority-queue methods. Does not achieve O(n log n) complexity of D8. More complex to implement.',
        'Does not incorporate flood prediction into routing decisions. Focuses on stochastic travel times but not environmental hazard prediction. No integration with real-time flood propagation models.',
        '30m resolution overestimates flood extent in some cases. Resolution mismatch issues with OSMnx roads. Suggests higher resolution (Copernicus GLO-30 at 30cm) would be better. May miss micro-topographic features.',
        'Focuses on EMS supply-demand mismatch but not predictive dispatch. Does not provide algorithmic solutions, only problem analysis. Limited to post-disaster analysis. No integration with flood propagation modeling.',
        'Requires extensive training data (historical flood events). Black-box nature makes predictions difficult to interpret. May not generalize to unprecedented flood scenarios. Our D8 approach is more interpretable.',
        'Focuses on pre-disaster positioning, not dynamic repositioning during disaster. Does not incorporate real-time flood propagation. Greedy approximation ratio (1 - 1/e) is suboptimal. Simplified MCLP without capacity constraints.',
        'Focuses on pathfinding, not disaster response scenarios. Does not provide specific reward weights (we used heuristics). Suggests inverse RL for principled weight assignment (our future work). Weights are hand-tuned.'
    ]
}

# Create DataFrame
df = pd.DataFrame(data)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Literature Survey"

# Define styles
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # SRM Blue
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
cell_font = Font(name='Calibri', size=10)
wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
center_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Add title
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = "Literature Survey Table — Top 10 Most Relevant Papers"
title_cell.font = Font(name='Calibri', size=14, bold=True, color="1E3A8A")
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Add headers
headers = ['No.', 'Paper Title', 'Ref', 'Year', 'Venue', 'How We Used This Paper', 'Limitations Stated']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Add data
for row_num, row_data in enumerate(df.itertuples(index=False), 3):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = cell_font
        cell.border = border
        
        # Apply wrap text for long columns
        if col_num in [2, 6, 7]:  # Paper Title, How We Used, Limitations
            cell.alignment = wrap_alignment
        elif col_num in [1, 3, 4]:  # No., Ref, Year
            cell.alignment = center_alignment
        else:
            cell.alignment = Alignment(vertical='top', horizontal='left')

# Set column widths
ws.column_dimensions['A'].width = 5   # No.
ws.column_dimensions['B'].width = 45  # Paper Title
ws.column_dimensions['C'].width = 6   # Ref
ws.column_dimensions['D'].width = 6   # Year
ws.column_dimensions['E'].width = 25  # Venue
ws.column_dimensions['F'].width = 50  # How We Used
ws.column_dimensions['G'].width = 50  # Limitations

# Set row heights
ws.row_dimensions[1].height = 25  # Title row
ws.row_dimensions[2].height = 30  # Header row
for row in range(3, 13):
    ws.row_dimensions[row].height = 100  # Data rows

# Freeze panes (freeze header)
ws.freeze_panes = 'A3'

# Save the workbook
output_file = 'Literature_Survey_Table.xlsx'
wb.save(output_file)

print(f"✅ Excel file created successfully: {output_file}")
print(f"\n📊 Table contains {len(df)} papers")
print(f"\n📝 Columns:")
for i, col in enumerate(headers, 1):
    print(f"   {i}. {col}")
print(f"\n💡 Tips:")
print(f"   • Open in Excel/PowerPoint")
print(f"   • Copy entire table and paste into PowerPoint")
print(f"   • Or insert as Excel object for live updates")
print(f"   • Column widths are pre-optimized for readability")
print(f"   • Text wrapping is enabled for long content")

# Also save as CSV for compatibility
df.to_csv('Literature_Survey_Table.csv', index=False)
print(f"\n✅ CSV file also created: Literature_Survey_Table.csv")
